from flask import Blueprint, request, jsonify, send_file
from models import db, GSTBill, GSTBillItem, GSTCustomer, Settings
from datetime import datetime
import csv
import io
import zipfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

gst_bill_bp = Blueprint('gst_bills', __name__)

def number_to_words(n):
    """Convert number to words in Indian format"""
    def convert_hundreds(num):
        ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
                'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
                'Seventeen', 'Eighteen', 'Nineteen']
        tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
        
        result = ''
        if num >= 100:
            result += ones[num // 100] + ' Hundred '
            num %= 100
        if num >= 20:
            result += tens[num // 10] + ' '
            num %= 10
        if num > 0:
            result += ones[num] + ' '
        return result

    if n == 0:
        return 'Zero Rupees Only'
    
    crores = n // 10000000
    n %= 10000000
    lakhs = n // 100000
    n %= 100000
    thousands = n // 1000
    n %= 1000
    hundreds = n
    
    result = ''
    if crores > 0:
        result += convert_hundreds(crores) + 'Crore '
    if lakhs > 0:
        result += convert_hundreds(lakhs) + 'Lakh '
    if thousands > 0:
        result += convert_hundreds(thousands) + 'Thousand '
    if hundreds > 0:
        result += convert_hundreds(hundreds)
    
    return result.strip() + ' Rupees Only'

@gst_bill_bp.route('/gst-bills', methods=['GET'])
def get_all_gst_bills():
    try:
        bills = GSTBill.get_all()
        return jsonify([bill.to_dict() for bill in bills])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@gst_bill_bp.route('/gst-bills/<int:bill_id>', methods=['GET'])
def get_gst_bill(bill_id):
    try:
        bill = GSTBill.get_by_id(bill_id)
        if not bill:
            return jsonify({'error': 'GST bill not found'}), 404
        return jsonify(bill.to_dict())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@gst_bill_bp.route('/gst-bills', methods=['POST'])
def create_gst_bill():
    try:
        data = request.get_json()
        
        # Convert date string to date object
        if isinstance(data.get('date'), str):
            data['date'] = datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        # Convert amount to words
        grand_total = float(data.get('grand_total', 0))
        data['amount_in_words'] = number_to_words(int(grand_total))
        
        # Store customer data
        if data.get('customer_name'):
            GSTCustomer.create_or_update(
                customer_name=data['customer_name'],
                customer_address=data.get('customer_address'),
                customer_gstin=data.get('customer_gstin')
            )
        
        # Create GST bill
        items_data = data.pop('items', [])
        bill = GSTBill.create(**data)
        
        # Create GST bill items
        for item_data in items_data:
            item_data['gst_bill_id'] = bill.id
            GSTBillItem.create(**item_data)
        
        db.session.commit()
        return jsonify(bill.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    # except Exception as e:
    #     db.session.rollback()
    #     return jsonify({'error': str(e)}), 500

@gst_bill_bp.route('/gst-bills/<int:bill_id>', methods=['PUT'])
def update_gst_bill(bill_id):
    try:
        bill = GSTBill.get_by_id(bill_id)
        if not bill:
            return jsonify({'error': 'GST bill not found'}), 404

        data = request.get_json()

        # Remove timestamps from client
        data.pop('created_at', None)
        data.pop('updated_at', None)

        # Parse 'date' if needed
        if 'date' in data and isinstance(data['date'], str):
            data['date'] = datetime.strptime(data['date'], '%Y-%m-%d').date()

        # Convert grand_total → words
        if 'grand_total' in data:
            grand_total = float(data['grand_total'])
            data['amount_in_words'] = number_to_words(int(grand_total))

        # Update bill fields
        items_data = data.pop('items', None)
        for key, value in data.items():
            if hasattr(bill, key):
                setattr(bill, key, value)

        # Replace items if provided
        if items_data is not None:
            GSTBillItem.query.filter_by(gst_bill_id=bill.id).delete()
            for item_data in items_data:
                item_data.pop('created_at', None)   # avoid error in items too
                item_data.pop('updated_at', None)
                item_data['gst_bill_id'] = bill.id
                db.session.add(GSTBillItem(**item_data))

        # Always set updated_at
        bill.updated_at = datetime.utcnow()

        db.session.commit()
        return jsonify(bill.to_dict()), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@gst_bill_bp.route('/gst-bills/<int:bill_id>/delete', methods=['POST'])
def delete_gst_bill(bill_id):
    try:
        bill = GSTBill.get_by_id(bill_id)
        if not bill:
            return jsonify({'error': 'GST bill not found'}), 404
        
        db.session.delete(bill)
        db.session.commit()
        return jsonify({'message': 'GST bill deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@gst_bill_bp.route('/gst-bills/export', methods=['GET'])
def export_gst_bills_excel():
    try:
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        customer_name = request.args.get('customer_name')
        
        # Convert date strings to date objects
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
        
        # Get filtered bills
        bills = GSTBill.get_filtered(start_date_obj, end_date_obj, customer_name)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "GST Bills"
        
        # Headers
        headers = [
            'Bill Number', 'Date', 'Customer Name', 'Customer Address', 'Customer GSTIN',
            'Total Before Tax', 'CGST %', 'SGST %', 'IGST %', 
            'CGST Amount', 'SGST Amount', 'IGST Amount', 'Grand Total'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, bill in enumerate(bills, 2):
            ws.cell(row=row, column=1, value=bill.bill_number)
            ws.cell(row=row, column=2, value=bill.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=3, value=bill.customer_name)
            ws.cell(row=row, column=4, value=bill.customer_address)
            ws.cell(row=row, column=5, value=bill.customer_gstin or '')
            ws.cell(row=row, column=6, value=bill.total_amount_before_tax)
            ws.cell(row=row, column=7, value=bill.cgst_percentage)
            ws.cell(row=row, column=8, value=bill.sgst_percentage)
            ws.cell(row=row, column=9, value=bill.igst_percentage)
            ws.cell(row=row, column=10, value=bill.cgst_amount)
            ws.cell(row=row, column=11, value=bill.sgst_amount)
            ws.cell(row=row, column=12, value=bill.igst_amount)
            ws.cell(row=row, column=13, value=bill.grand_total)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'gst_bills_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@gst_bill_bp.route('/gst-bills/export-csv', methods=['GET'])
def export_gst_bills_csv():
    try:
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        customer_name = request.args.get('customer_name')
        
        # Convert date strings to date objects
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
        
        # Get filtered bills
        bills = GSTBill.get_filtered(start_date_obj, end_date_obj, customer_name)
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            'Bill Number', 'Date', 'Customer Name', 'Customer Address', 'Customer GSTIN',
            'Total Before Tax', 'CGST %', 'SGST %', 'IGST %', 
            'CGST Amount', 'SGST Amount', 'IGST Amount', 'Grand Total'
        ])
        
        # Data rows
        for bill in bills:
            writer.writerow([
                bill.bill_number,
                bill.date.strftime('%Y-%m-%d'),
                bill.customer_name,
                bill.customer_address,
                bill.customer_gstin or '',
                bill.total_amount_before_tax,
                bill.cgst_percentage,
                bill.sgst_percentage,
                bill.igst_percentage,
                bill.cgst_amount,
                bill.sgst_amount,
                bill.igst_amount,
                bill.grand_total
            ])
        
        # Convert to bytes
        output.seek(0)
        csv_data = output.getvalue().encode('utf-8')
        
        filename = f'gst_bills_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return send_file(
            io.BytesIO(csv_data),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def get_gst_bills_csv_data():
    """Helper function to get GST bills data as CSV for backup"""
    try:
        bills = GSTBill.get_all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow([
            'Bill Number', 'Date', 'Customer Name', 'Customer Address', 'Customer GSTIN',
            'Items', 'Total Before Tax', 'CGST %', 'SGST %', 'IGST %', 
            'CGST Amount', 'SGST Amount', 'IGST Amount', 'Grand Total', 'Amount in Words'
        ])
        
        # Data rows
        for bill in bills:
            items_str = '; '.join([f"{item.description} (HSN: {item.hsn}, Weight: {item.weight}, Rate: {item.rate}, Amount: {item.amount})" for item in bill.items])
            writer.writerow([
                bill.bill_number,
                bill.date.strftime('%Y-%m-%d'),
                bill.customer_name,
                bill.customer_address,
                bill.customer_gstin or '',
                items_str,
                bill.total_amount_before_tax,
                bill.cgst_percentage,
                bill.sgst_percentage,
                bill.igst_percentage,
                bill.cgst_amount,
                bill.sgst_amount,
                bill.igst_amount,
                bill.grand_total,
                bill.amount_in_words
            ])
        
        return output.getvalue()
    except Exception as e:
        print(f"Error generating GST bills CSV: {e}")
        return ""